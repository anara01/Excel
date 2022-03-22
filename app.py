#from flask import Flask, render_template, request
#import pandas as pd
#from openpyxl import Workbook, load_workbook


#app = Flask(__name__)

#@app.route('/')
#def data():
#	data = pd.read_excel('good.xlsx')
#	return render_template('index.html', data=data)


#@app.route('/add/', methods=["POST"])
#def add():
#	good = request.form["good"]

#	excel_file = Workbook()
#	excel_sheet = excel_file.create_sheet(title='Sheet1', index=0)

#	excel_sheet.append([good])

#	excel_file.save(filename='good.xlsx')

#	return """
#		<h1>Инвентарь пополнен</h1>
#		<a href='/'>Домой</a>
#	"""

#___________________________


from flask import Flask, render_template, request
import pandas as pd
import openpyxl
app = Flask(__name__)
book = openpyxl.open('goods.xlsx', read_only=True)
sheet = book.active
goods = []
i = 1
while sheet[f'A{i}'].value:
    goods.append(sheet[f'A{i}'].value)
    #sheet[f'A{i}'].value = goods
    i += 1
book.close()
#book.close()
#return render_template('index.html', goods=goods)
@app.route('/')
def homepage():
    book = openpyxl.open('goods.xlsx', read_only=True)
    #sheet = book.active
	#book = pd.read_excel('goods.xlsx')

    #i = 1
    #while sheet[f'A{i}'].value:
    #    goods.append(sheet[f'A{i}'].value)
    #    #sheet[f'A{i}'].value = goods
    #    i += 1
    #book.close()
    return render_template('index.html', goods=goods)
@app.route('/add/', methods=["POST"])
def add():
    good = request.form["good"]
    goods.append(good)
    book = openpyxl.open('goods.xlsx')
    sheet = book.active
    x = len(goods)
    sheet[x][0].value = good
    book.save('goods.xlsx')
    book.close()
    return """
        <h1>Инвентарь пополнен</h1>
        <a href='/'>Домой</a>
    """
if __name__ == '__main__':
    app.run(debug=True)
